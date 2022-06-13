# import領域
import openpyxl
import datetime

# 変数のglobal化
sample_number = 1
cont_name = ""
cont_mw = 0
cont_concentration = 0
substrates = ["HTF", "HLF", "rHLF", "BLF", "OVA", "CFT (single use only)", "Human-IgG", "Sheep-IgG", "Taka-Amylase", "alpha1-AG", "RNaseB"]
sample_areas = []

# ファイルの新規作成をする場合
print("create new file?")
mode_file = int(input("0:new file, 1:open file, selected :"))
if mode_file == 0:
    dt_now = datetime.datetime.now()
    file_number = input("what sheets do you create today? :")
    created_filename = str(dt_now.year) + str(dt_now.month) + str(dt_now.day) + "_" + file_number + ".xlsx"
    wb = openpyxl.Workbook("../desktop/" + created_filename)
    filename = "../desktop/" + created_filename
    wb.save(filename)
# ファイルの指定をする場合
else:
    selected_workbook = input("enter workbook_name :")
    filename = selected_workbook

wb = openpyxl.load_workbook(filename)
ws = wb.active

# tableの用意
ws.cell(3, 1).value = "反応時間"
ws.cell(4, 1).value = "希釈倍率"
ws.cell(5, 1).value = "積算値"
ws.cell(11, 1).value = "割合"
ws.cell(17, 1).value = "U1/ml(*10-5)"
ws.cell(18, 1).value = "U2/ml"
ws.cell(5, 2).value = "S"
ws.cell(11, 2).value = "S"
for i in range(5):
    num = "P" + str(i + 1)
    ws.cell(6 + i, 2).value = num
    ws.cell(12 + i, 2).value = num
ws.column_dimensions["A"].width = 9
ws.column_dimensions["B"].width = 3
wb.save(filename)

# 規定条件
reaction_time_ori = int(input("reaction time? :"))
dilution_ori = int(input("dilution? :"))
exit_ = "c"
while True:
    # 基質の設定
    if exit_ == "c":
        for i in range(len(substrates)):
            print(str(i) + ":" + substrates[i])
        wb_substrates = openpyxl.load_workbook("substrates.xlsx")
        ws_substrates = wb_substrates.active
        print("**" + str(sample_number) + "**  ~cont.")
        control_substrate = int(input("what substrate?(int only input)"))
        cont_name = ws_substrates.cell(row = control_substrate + 1, column = 1).value
        cont_mw = ws_substrates.cell(row = control_substrate + 1, column = 3).value
        cont_concentration = ws_substrates.cell(row = control_substrate + 1, column = 2).value
        wb_substrates.close()
        # cont.の設定
        area_list = []
        cont_ratio_list = []
        condition = input("what condition? 0:origin, 1:unique")
        if condition == "0":
            reaction_time = reaction_time_ori
            dilution = dilution_ori
        
        else:
            reaction_time = int(input("reaction time? :"))
            dilution = int(input("dilution? :"))

        area_list.append(int(input("S: ")))
        for i in range(5):
            area_var = input("P" + str(i + 1) + " :")
            if area_var:
                area_list.append(int(area_var))
            else:
                area_list.append(0)
        sum_ = 0
        for i in range(6):
            ws.cell(5 + i, sample_number + 2).value = area_list[i]
            sum_ += area_list[i]
        for i in range(6):
            cont_ratio_list.append(float(area_list[i] / sum_))
            ws.cell(11 + i, sample_number + 2).value = cont_ratio_list[i]
        ws.cell(2, sample_number + 2).value = cont_name
        ws.cell(3, sample_number + 2).value = reaction_time
        sample_number += 1
        # サンプル入力
        print("**" + str(sample_number) + "**")
        area_list = []
        ratio_list = []
        condition = input("what condition? 0:origin, 1:unique")
        if condition == "0":
            reaction_time = reaction_time_ori
            dilution = dilution_ori
        else:
            reaction_time = int(input("reaction time? :"))
            dilution = int(input("dilution? :"))
        area_list.append(int(input("S: ")))
        for i in range(5):
            area_var = input("P" + str(i + 1) + " :")
            if area_var:
                area_list.append(int(area_var))
            else:
                area_list.append(0)
        sum_ = 0
        for i in range(6):
            ws.cell(5 + i, sample_number + 2).value = area_list[i]
            sum_ += area_list[i]
        for i in range(6):
            ratio_list.append(float(area_list[i] / sum_))
            ws.cell(11 + i, sample_number + 2).value = ratio_list[i]
        sum_glycan = 0
        for i in range(5):
            sum_glycan += ratio_list[i + 1] * (i + 1)
            sum_glycan -= cont_ratio_list[i + 1] * (i + 1)
        ws.cell(3, sample_number + 2).value = reaction_time
        ws.cell(4, sample_number + 2).value = dilution
        ws.cell(17, sample_number + 2).value = float(sum_glycan * cont_concentration * dilution * 5000000 / 3 / reaction_time / cont_mw)
        ws.cell(18, sample_number + 2).value = float((1 - ratio_list[0] - cont_ratio_list[1] - cont_ratio_list[2]) * cont_concentration * dilution * 50 / 3 / reaction_time)
        sample_number += 1
        exit_ = ""
        print("******")
        exit_ = input("c: control, e:exit, selected :")
        if exit_ == "e":
            break
    else:
        # サンプル入力

        print("**" + str(sample_number) + "**")
        area_list = []
        ratio_list = []
        condition = input("what condition? 0:origin, 1:unique")
        if condition == "0":
            reaction_time = reaction_time_ori
            dilution = dilution_ori
        else:
            reaction_time = int(input("reaction time? :"))
            dilution = int(input("dilution? :"))

        area_list.append(int(input("S: ")))
        for i in range(5):
            area_var = input("P" + str(i + 1) + " :")
            if area_var:
                area_list.append(int(area_var))
            else:
                area_list.append(0)
        sum_ = 0
        for i in range(6):
            ws.cell(5 + i, sample_number + 2).value = area_list[i]
            sum_ += area_list[i]
        for i in range(6):
            ratio_list.append(float(area_list[i] / sum_))
            ws.cell(11 + i, sample_number + 2).value = ratio_list[i]
        sum_glycan = 0
        for i in range(5):
            sum_glycan += ratio_list[i + 1] * (i + 1)
            sum_glycan -= cont_ratio_list[i + 1] * (i + 1)
        ws.cell(3, sample_number + 2).value = reaction_time
        ws.cell(4, sample_number + 2).value = dilution
        ws.cell(17, sample_number + 2).value = float(sum_glycan * cont_concentration * dilution * 5000000 / 3 / reaction_time / cont_mw)
        ws.cell(18, sample_number + 2).value = float((1 - ratio_list[0] - cont_ratio_list[1] - cont_ratio_list[2]) * cont_concentration * dilution * 50 / 3 / reaction_time)
        sample_number += 1
        exit_ = ""
        print("******")
        exit_ = input("c: control, e:exit, selected :")
        if exit_ == "e":
            break

# 最終調整
if control_substrate != 5:
    ws.delete_rows(8, 3)
    ws.delete_rows(11, 3)

wb.save(filename)
wb.close()

print("******")